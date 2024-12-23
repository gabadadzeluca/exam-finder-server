from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2 import service_account
import openpyxl
import io
import os
import json

app = Flask(__name__)
# cors = CORS(app, origins="https://freeuni-examfinder.netlify.app")
CORS(app, resources={r"/api/*": {"origins": "*"}, r"/excel/download*": {"origins": "https://freeuni-examfinder.netlify.app"}})

# Load the JSON content from the environment variable
service_account_info = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')

# Convert the JSON string back to a dictionary
service_account_data = json.loads(service_account_info)

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]

creds = service_account.Credentials.from_service_account_info(
        service_account_data, scopes=SCOPES)

SPREADSHEET_ID = "1eUABgVBLCX82ixJ8eCSqAm6hflPWZucKe5-rjcx9CsA" #2025 I semester docs
RANGE_NAME = "A1:F20" # Range of needed data;

exams = []

@app.route("/api/data", methods=["GET"])
def data():
    uniGroup = request.args.get('uniGroup')
    examData = getExamData(uniGroup)
    return jsonify(
        {
          "examData": examData
        }
    )

@app.route("/excel/download", methods=["GET", "POST"])
def download_excel():
  # Check if it's a POST request and read from JSON body
  if request.method == "POST":
      data = request.get_json().get('examData')
  else:  # For GET requests, keep using query parameters
      data = json.loads(request.args.get('examData'))
  output = excelHandler(data)
  return send_file(output, 
          as_attachment=True, 
          download_name='exams_excel.xlsx', 
          mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def getExamData(uniGroup):
  exams = []; # intilize again to avoid accumulating data
  try:
    service = build("sheets", "v4", credentials=creds)

    spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID, fields="sheets(properties(title))").execute()

    # Get the list of sheets
    sheets = spreadsheet.get('sheets', [])

    ranges = [f"{sheetObj['properties']['title']}!{RANGE_NAME}" for sheetObj in sheets]

    # Call the Sheets API using batchGet
    result = service.spreadsheets().values().batchGet(spreadsheetId=SPREADSHEET_ID, ranges=ranges).execute()

    # Process the results
    value_ranges = result.get('valueRanges', [])
    for value_range in value_ranges:
      sheet_title = value_range['range'].split('!')[0]
      print(sheet_title)
      values = value_range.get('values', [])
      for row in values:
        if len(row) > 3: # Check for invalid sheet, as in tests
          if uniGroup in row[3]:
              # append title (date) to the info; title is formatted as a string, 
              # e.g. sheet_title = '18/06'; saving it as 18/06 to later take month and day as separate variables 
              row.append(sheet_title[1:len(sheet_title)-1])
              exams.append(row)
        else:
          print(f"Row in the sheet {sheet_title} is too short")
    print(exams)
    return exams

  except HttpError as err:
    print(err)

def excelHandler(data):
  output = io.BytesIO()
  
  try:
    # Try to load an existing workbook
    wb = openpyxl.load_workbook('exams_excel.xlsx')
    print("Workbook exists.")
  except FileNotFoundError:
    # If doesn't exist, create a new one
    print("Workbook doesn't exist. Creating a new workbook.")
    wb = openpyxl.Workbook()

  # Select the active sheet
  sheet = wb.active

  for row_idx, row in enumerate(data):
    for col_idx, cell_value in enumerate(row):
      print(col_idx, cell_value)
      if col_idx == 6:
        # date format is day/month
        day, month = cell_value.split('/')
        sheet.cell(row=row_idx + 1, column=col_idx + 1, value=f"=DATE(2024,{month},{day})").number_format = "DD/MM/YYYY"
      else:
        sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

  # Save the workbook
  wb.save(output)
  output.seek(0)  # Rewind the buffer to the beginning
  return output


if __name__ == "__main__":
  app.run(debug=True, port=5000)

